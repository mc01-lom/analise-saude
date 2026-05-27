"""
pipeline_rfi_to_painel.py
─────────────────────────────────────────────────────────────────────────────
Lê a aba 'DRE BPJ' de cada RFI de fornecedor e alimenta a seção
'Abertura de Custos' do Painel_BUPJ.xlsx com os dados mensais.

Configuração: edite o arquivo config.xlsx (abas Geral e Fornecedores).
Uso         : python pipeline_rfi_to_painel.py
─────────────────────────────────────────────────────────────────────────────
"""

from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ═══════════════════════════════════════════════════════════════════════════
# 1. LEITURA DO CONFIG
# ═══════════════════════════════════════════════════════════════════════════

CONFIG_FILE = Path(__file__).parent / "config.xlsx"
DRE_TAB     = "DRE BPJ"


def load_config() -> tuple[Path, Path, dict]:
    """
    Lê config.xlsx e retorna:
        painel_file : Path
        output_file : Path
        suppliers   : { nome: { ano: (col_inicio, col_fim) } }
                      colunas em inteiros 1-indexed
    """
    wb  = load_workbook(CONFIG_FILE, data_only=True)

    # ── Aba Geral ────────────────────────────────────────────────────────
    ws = wb["Geral"]
    geral = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            geral[row[0].strip()] = str(row[1]).strip()

    base    = Path(__file__).parent
    painel  = base / geral.get("Arquivo Painel", "Painel_BUPJ.xlsx")
    output  = base / geral.get("Arquivo Saída",  "Painel_BUPJ_atualizado.xlsx")

    # ── Aba Fornecedores ─────────────────────────────────────────────────
    ws = wb["Fornecedores"]
    suppliers: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        nome, rfi_file, ano, col_ini, col_fim = (row + (None,) * 5)[:5]
        if not all([nome, rfi_file, ano, col_ini, col_fim]):
            continue
        # Ignora linhas de instrução/nota (col A não é fornecedor)
        if str(nome).startswith("ℹ"):
            continue

        nome     = str(nome).strip()
        rfi_file = base / str(rfi_file).strip()
        ano      = int(ano)
        col_ini  = column_index_from_string(str(col_ini).strip())
        col_fim  = column_index_from_string(str(col_fim).strip())

        if nome not in suppliers:
            suppliers[nome] = {"rfi_file": rfi_file, "anos": {}}
        suppliers[nome]["anos"][ano] = (col_ini, col_fim)

    # ── Aba Mapeamento ───────────────────────────────────────────────────
    ws = wb["Mapeamento"]
    mapping_rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        nome_dre, nome_painel, bloco, linha = (list(row) + [None]*4)[:4]
        # Pula linhas vazias ou de nota
        if not nome_painel or str(nome_painel).startswith("ℹ"):
            continue
        mapping_rows.append({
            "nome_dre"   : str(nome_dre).strip()   if nome_dre   else None,
            "nome_painel": str(nome_painel).strip() if nome_painel else None,
            "bloco"      : str(bloco).strip()       if bloco      else None,
            "linha"      : int(linha)               if linha      else None,
        })

    wb.close()
    return painel, output, suppliers, mapping_rows


# ═══════════════════════════════════════════════════════════════════════════
# 2. CONSTRUÇÃO DINÂMICA DO MAPEAMENTO (a partir do config)
# ═══════════════════════════════════════════════════════════════════════════


# Grupos da DRE com suas linhas de total e catch-all no Painel
# catch_all=None → itens "outros" desse grupo são descartados
SECTIONS: list[dict] = [
    {"key": "despesas diretas com pessoal",        "header_row": 75,  "catch_all": 87},
    {"key": "despesas com serviços especializados", "header_row": 94,  "catch_all": 114},
    {"key": "ações digitais",                       "header_row": 117, "catch_all": None},
    {"key": "infraestrutura",                       "header_row": 140, "catch_all": 154},
    {"key": "outras despesas",                      "header_row": 160, "catch_all": 172},
    {"key": "impostos",                             "header_row": 182, "catch_all": None},
]

def build_mapping(mapping_rows: list[dict]) -> tuple[dict, dict, tuple, dict, set, set, set]:
    """
    Constrói as estruturas de mapeamento a partir das linhas lidas do config.

    Retorna:
        label_to_row        : { label_lower → linha }     — mapeamento direto
        multi_target        : { label_lower → [l1, l2] }  — multi-destino
        outros_prefixes     : tuple de prefixos CURTOS para detecção de "outros"
        outros_map          : { prefixo_curto → linha | None }  — destino direto
        skip_labels         : labels a descartar (fora do escopo)
        discard_labels      : labels sem destino (linha=None no config)
        section_header_exact: headers exatos de seção na DRE
    """
    label_to_row: dict[str, int]        = {}
    multi_target: dict[str, list[int]]  = {}
    discard_labels: set[str]            = set()

    skip_labels: set[str] = {
        "receita", "receita fixa", "receita variável", "resultado",
    }

    # Prefixos CURTOS estáveis — tolerantes a variações singular/plural da DRE
    # Mapeados para linha de destino (None = descartar)
    outros_map: dict[str, int | None] = {
        "outras despesas direcionada a pessoas"          : 87,
        "outras despesas direcionadas a pessoas"         : 87,
        "outros serviços especializados 2"               : 114,
        "outros serviços especializados 3"               : 114,
        "outros serviços especializados 4"               : 114,
        "outros serviços especializados 5"               : 114,
        "outras despesas direcionadas a infraestrutura"  : 154,
        "outras despesas direcionada a infraestrutura"   : 154,
        "outras despesas 1"                              : 172,
        "outras despesas 2"                              : 172,
        "outras despesas 3"                              : 172,
        "outras despesas 4"                              : 172,
        "outras despesas 5"                              : 172,
        "outras despesas 6"                              : 172,
        "outras despesas 7"                              : 172,
        "outras despesas 8"                              : 172,
        "outras despesas 9"                              : 172,
        "outras despesas 10"                             : 172,
        "outras despesas direcionadas a ações digitais"  : None,
        "outras despesas direcionada a ações digitais"   : None,
    }
    outros_prefixes = tuple(outros_map.keys())

    for row in mapping_rows:
        dre   = row["nome_dre"]
        linha = row["linha"]

        if not dre:
            continue   # sem correspondência DRE → nada a mapear

        label = dre.strip().lower()

        if linha is None:
            discard_labels.add(label)
            continue

        # "outros" já tratados via outros_map — pular
        if any(label.startswith(p) for p in outros_prefixes):
            continue

        # Mapeamento normal ou multi-target
        if label in label_to_row and label_to_row[label] != linha:
            existing = label_to_row.pop(label)
            multi_target[label] = sorted({existing, linha})
        elif label in multi_target:
            if linha not in multi_target[label]:
                multi_target[label] = sorted(set(multi_target[label]) | {linha})
        else:
            label_to_row[label] = linha

    section_header_exact = {
        "despesas diretas com pessoal",
        "despesas com serviços especializados (consultoria, auditoria, monitoria, informações cadastrais, etc)",
        "ações digitais",
        "infraestrutura",
        "outras despesas",
        "impostos",
    }

    return label_to_row, multi_target, outros_prefixes, outros_map, skip_labels, discard_labels, section_header_exact

# ═══════════════════════════════════════════════════════════════════════════
# 2. MAPEAMENTO DRE → PAINEL
# ═══════════════════════════════════════════════════════════════════════════

# Preenchidos em runtime a partir do config.xlsx via build_mapping()
LABEL_TO_ROW         : dict[str, int]       = {}
MULTI_TARGET_ROWS    : dict[str, list[int]] = {}
OUTROS_PREFIXES      : tuple[str, ...]      = ()
OUTROS_MAP           : dict[str, int | None]= {}   # prefixo → linha destino
SKIP_LABELS          : set[str]             = set()
_DISCARD_LABELS      : set[str]             = set()
SECTION_HEADER_EXACT : set[str]             = set()


# ═══════════════════════════════════════════════════════════════════════════
# 3. FUNÇÕES AUXILIARES
# ═══════════════════════════════════════════════════════════════════════════

def normalize(text: str) -> str:
    return text.strip().lower() if isinstance(text, str) else ""


def current_section_info(label_lower: str) -> dict | None:
    if any(label_lower.startswith(p) for p in OUTROS_PREFIXES):
        return None
    for sec in SECTIONS:
        if label_lower.startswith(sec["key"]):
            return sec
    return None


def resolve_painel_row(label_lower: str, current_sec: dict | None) -> int | list[int] | None:
    """
    Retorna:
        int       → linha única do Painel
        list[int] → múltiplas linhas (multi-target)
        -1        → catch-all da seção corrente
        None      → descartar

    Ordem de prioridade:
        1. SKIP_LABELS          → descartar
        2. OUTROS_PREFIXES      → catch-all
        3. MULTI_TARGET_ROWS    → múltiplos destinos
        4. LABEL_TO_ROW         → destino direto  ← antes de seção para evitar
                                                      conflitos de startswith
                                                      (ex: "impostos e taxas"
                                                       vs seção "impostos")
        5. Header de seção      → linha do total do grupo
        6. Sem mapeamento + seção com catch-all → catch-all
        7. Sem mapeamento       → descartar
    """
    if label_lower in SKIP_LABELS or label_lower in _DISCARD_LABELS:
        return None

    # "Outros" → destino direto via OUTROS_MAP (evita dependência de current_sec)
    for prefix, dest_row in OUTROS_MAP.items():
        if label_lower.startswith(prefix):
            return dest_row  # None = descartar, int = somar nessa linha

    if label_lower in MULTI_TARGET_ROWS:
        return MULTI_TARGET_ROWS[label_lower]

    if label_lower in LABEL_TO_ROW:
        return LABEL_TO_ROW[label_lower]

    sec = current_section_info(label_lower)
    if sec:
        return sec["header_row"]

    if current_sec and current_sec["catch_all"]:
        return -1

    return None  # descartar


# ═══════════════════════════════════════════════════════════════════════════
# 4. LEITURA DA DRE
# ═══════════════════════════════════════════════════════════════════════════

def read_dre(rfi_path: Path, anos_desejados: list[int]) -> dict[int, dict[int, list[float]]]:
    """
    Lê a aba DRE BPJ e retorna:
        { painel_row: { ano: [v_jan, v_fev, ..., v_dez] } }

    Detecta dinamicamente quais colunas da DRE correspondem a cada ano.
    Processa apenas os anos presentes em anos_desejados.
    """
    print(f"    Abrindo aba '{DRE_TAB}'...")
    try:
        wb = load_workbook(rfi_path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(
            f"Não foi possível abrir '{rfi_path.name}': {e}\n"
            f"    Verifique se o arquivo é um .xlsx válido e não está corrompido."
        ) from e

    ws   = wb[DRE_TAB]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    print(f"    {len(rows)} linhas lidas da aba.")

    # ── Detectar colunas por ano (linha 3, índice 2) ─────────────────────
    # DRE: col B = label (idx 1), dados a partir de col C (idx 2)
    import datetime
    ano_to_cols: dict[int, list[int]] = {}   # ano → [idx_col, ...]

    header_row = rows[2]   # linha 3 (0-indexed: 2)
    for idx, cell in enumerate(header_row):
        if isinstance(cell, datetime.datetime):
            ano = cell.year
            if ano in anos_desejados:
                ano_to_cols.setdefault(ano, []).append(idx)

    for ano, cols in sorted(ano_to_cols.items()):
        print(f"    Ano {ano}: {len(cols)} meses detectados "
              f"(colunas DRE {cols[0]+1}–{cols[-1]+1})")

    anos_ausentes = [a for a in anos_desejados if a not in ano_to_cols]
    if anos_ausentes:
        print(f"    [AVISO] Anos não encontrados na DRE: {anos_ausentes}")

    # ── Processar linhas de dados ─────────────────────────────────────────
    result: dict[int, dict[int, list[float]]] = {}
    catch_all_acc: dict[int, dict[int, list[float]]] = {}   # {ca_row: {ano: [vals]}}
    current_sec: dict | None = None
    stats = {"mapeadas": 0, "catch_all": 0, "descartadas": 0}

    for row in rows[3:]:
        label_raw = row[1]
        if label_raw is None:
            continue
        label_lower = normalize(label_raw)
        if not label_lower:
            continue

        # Atualiza seção corrente apenas para labels exatos de header de grupo
        if label_lower in SECTION_HEADER_EXACT:
            new_sec = current_section_info(label_lower)
            if new_sec:
                current_sec = new_sec
                print(f"\n    [Seção] {label_raw}")

        dest = resolve_painel_row(label_lower, current_sec)

        if dest is None:
            stats["descartadas"] += 1
            continue

        # Extrai valores por ano
        valores_por_ano: dict[int, list[float]] = {}
        for ano, cols in ano_to_cols.items():
            vals = []
            for idx in cols:
                raw = row[idx] if idx < len(row) else None
                vals.append(float(raw) if isinstance(raw, (int, float)) else 0.0)
            valores_por_ano[ano] = vals

        if dest is None:
            stats["descartadas"] += 1
            print(f"      ↳ descartado : {label_raw}")
        elif dest == -1:
            # Fallback catch-all (item nomeado sem mapeamento dentro de seção)
            if current_sec and current_sec["catch_all"]:
                ca = current_sec["catch_all"]
                if ca not in catch_all_acc:
                    catch_all_acc[ca] = {}
                for ano, vals in valores_por_ano.items():
                    if ano not in catch_all_acc[ca]:
                        catch_all_acc[ca][ano] = [0.0] * len(vals)
                    for m, v in enumerate(vals):
                        catch_all_acc[ca][ano][m] += v
                stats["catch_all"] += 1
                print(f"      ↳ catch-all  : {label_raw}")
            else:
                stats["descartadas"] += 1
                print(f"      ↳ descartado : {label_raw}")
        elif isinstance(dest, list):
            # Multi-target: grava o mesmo valor em múltiplas linhas do Painel
            for d in dest:
                result[d] = valores_por_ano
            stats["mapeadas"] += 1
            rows_str = ", ".join(f"L{d}" for d in dest)
            print(f"      → {rows_str} : {label_raw}")
        else:
            # Destino direto — se já existe (ex: múltiplos "outros" → mesma linha), soma
            if dest in result:
                for ano, vals in valores_por_ano.items():
                    if ano in result[dest]:
                        for m, v in enumerate(vals):
                            result[dest][ano][m] += v
                    else:
                        result[dest][ano] = vals
                stats["catch_all"] += 1
                print(f"      ↳ acumulado L{dest:<4} : {label_raw}")
            else:
                result[dest] = valores_por_ano
                stats["mapeadas"] += 1
                print(f"      → L{dest:<4} : {label_raw}")

    # Mescla catch-alls
    for ca_row, anos_dict in catch_all_acc.items():
        if ca_row not in result:
            result[ca_row] = {}
        for ano, vals in anos_dict.items():
            if ano not in result[ca_row]:
                result[ca_row][ano] = vals
            else:
                for m, v in enumerate(vals):
                    result[ca_row][ano][m] += v

    print(f"\n    Resumo da leitura:")
    print(f"      • Linhas mapeadas    : {stats['mapeadas']}")
    print(f"      • Linhas no catch-all: {stats['catch_all']}")
    print(f"      • Linhas descartadas : {stats['descartadas']}")

    return result


# ═══════════════════════════════════════════════════════════════════════════
# 5. ESCRITA NO PAINEL
# ═══════════════════════════════════════════════════════════════════════════

def clear_supplier(ws, anos_cols: dict[int, tuple[int, int]],
                   row_start: int = 73, row_end: int = 192) -> int:
    """
    Zera todas as células do fornecedor no Painel antes de gravar novos dados.
    Limpa as linhas row_start..row_end nas colunas configuradas para cada ano.
    Retorna o número de células zeradas.
    """
    zeradas = 0
    for ano, (start, end) in anos_cols.items():
        for row in range(row_start, row_end + 1):
            for col in range(start, end + 1):
                if ws.cell(row=row, column=col).value is not None:
                    ws.cell(row=row, column=col).value = None
                    zeradas += 1
    return zeradas


def write_to_painel(ws, dre_data: dict[int, dict[int, list[float]]],
                    anos_cols: dict[int, tuple[int, int]]) -> int:
    """
    Grava os dados no Painel.
        anos_cols : { ano: (col_inicio, col_fim) }  — 1-indexed, inclusive
    Retorna o número de linhas gravadas.
    """
    gravadas = 0
    for painel_row, anos_dict in dre_data.items():
        for ano, vals in anos_dict.items():
            if ano not in anos_cols:
                continue
            start, end = anos_cols[ano]
            for i, col in enumerate(range(start, end + 1)):
                if i < len(vals):
                    ws.cell(row=painel_row, column=col).value = vals[i]
        gravadas += 1
    return gravadas


# ═══════════════════════════════════════════════════════════════════════════
# 6. DESCOBERTA E CONFIRMAÇÃO DE RFIs
# ═══════════════════════════════════════════════════════════════════════════

def discover_rfis() -> list[Path]:
    return sorted(Path(__file__).parent.glob("RFI_*.xlsx"))


def confirm_rfis(rfis: list[Path], suppliers: dict) -> bool:
    rfi_names = {p.name for p in rfis}
    mapped_names = {cfg["rfi_file"].name for cfg in suppliers.values()}

    print("=" * 60)
    print("  RFIs encontrados na pasta:")
    print("=" * 60)
    if not rfis:
        print("  [NENHUM arquivo RFI_*.xlsx encontrado]")
        return False

    for i, rfi in enumerate(rfis, start=1):
        status = "" if rfi.name in mapped_names else "  ⚠ não mapeado no config"
        print(f"  {i}. {rfi.name}{status}")

    ausentes = [
        (nome, cfg["rfi_file"].name)
        for nome, cfg in suppliers.items()
        if not cfg["rfi_file"].exists()
    ]
    if ausentes:
        print()
        print("  [AVISO] Fornecedores no config sem RFI na pasta:")
        for nome, fname in ausentes:
            print(f"    • {nome} → {fname}")

    print("=" * 60)
    resposta = input("  Digite OK para processar ou qualquer outra tecla para cancelar: ").strip()
    return resposta.upper() == "OK"


# ═══════════════════════════════════════════════════════════════════════════
# 7. MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main() -> None:
    # ── Carregar config ──────────────────────────────────────────────────
    if not CONFIG_FILE.exists():
        print(f"[ERRO] Arquivo de configuração não encontrado: {CONFIG_FILE}")
        return

    painel_file, output_file, suppliers, mapping_rows = load_config()
    anos_globais = sorted({
        ano
        for cfg in suppliers.values()
        for ano in cfg["anos"]
    })
    print(f"Configuração carregada: {len(suppliers)} fornecedor(es), "
          f"anos: {anos_globais}, {len(mapping_rows)} linhas de mapeamento.")

    # Popular variáveis globais de mapeamento
    global LABEL_TO_ROW, MULTI_TARGET_ROWS, OUTROS_PREFIXES, OUTROS_MAP, SKIP_LABELS, _DISCARD_LABELS
    LABEL_TO_ROW, MULTI_TARGET_ROWS, OUTROS_PREFIXES, OUTROS_MAP, SKIP_LABELS, _DISCARD_LABELS, SECTION_HEADER_EXACT = \
        build_mapping(mapping_rows)
    print(f"  Mapeamento: {len(LABEL_TO_ROW)} diretos, "
          f"{len(MULTI_TARGET_ROWS)} multi-target, "
          f"{len(OUTROS_PREFIXES)} catch-all, "
          f"{len(_DISCARD_LABELS)} descartados.")

    # ── Descoberta e confirmação ─────────────────────────────────────────
    rfis = discover_rfis()
    if not confirm_rfis(rfis, suppliers):
        print("\nOperação cancelada pelo usuário.")
        return

    total = sum(1 for cfg in suppliers.values() if cfg["rfi_file"].exists())
    processados = 0

    # ── Carregar Painel ──────────────────────────────────────────────────
    print(f"\n[1/3] Carregando Painel base: {painel_file.name}...")
    if not painel_file.exists():
        print(f"[ERRO] Painel não encontrado: {painel_file}")
        return
    wb_painel = load_workbook(painel_file)
    ws_painel = wb_painel.active
    print(f"      Painel carregado.")

    # ── Processar fornecedores ───────────────────────────────────────────
    print(f"\n[2/3] Processando {total} fornecedor(es)...")

    for supplier, cfg in suppliers.items():
        rfi_path  = cfg["rfi_file"]
        anos_cols = cfg["anos"]   # { ano: (col_ini, col_fim) }

        if not rfi_path.exists():
            print(f"\n  [AVISO] {rfi_path.name} não encontrado — pulando {supplier}.")
            continue

        processados += 1
        anos_str = ", ".join(str(a) for a in sorted(anos_cols))
        print(f"\n  ({processados}/{total}) {supplier} — {rfi_path.name}")
        print(f"  Anos: {anos_str}")
        print(f"  {'─' * 55}")

        # ── Etapa A: leitura do RFI ──────────────────────────────────────
        print(f"\n    [Etapa A] Lendo RFI...")
        try:
            dre_data = read_dre(rfi_path, list(anos_cols.keys()))
        except ValueError as e:
            print(f"\n    [ERRO] {e}")
            print(f"    Pulando {supplier} e continuando...")
            continue

        # ── Etapa B: limpeza das colunas no Painel ───────────────────────
        print(f"\n    [Etapa B] Limpando dados anteriores no Painel...")
        for ano, (start, end) in sorted(anos_cols.items()):
            from openpyxl.utils import get_column_letter
            print(f"      {ano}: colunas {get_column_letter(start)}–{get_column_letter(end)}")
        zeradas = clear_supplier(ws_painel, anos_cols)
        print(f"      {zeradas} células zeradas.")

        # ── Etapa C: gravação dos novos dados ────────────────────────────
        print(f"\n    [Etapa C] Gravando novos dados no Painel...")
        gravadas = write_to_painel(ws_painel, dre_data, anos_cols)
        print(f"      {gravadas} linhas gravadas com sucesso.")

    # ── Salvar ───────────────────────────────────────────────────────────
    print(f"\n[3/3] Salvando: {output_file.name}...")
    wb_painel.save(output_file)
    print(f"      Arquivo salvo.")

    print(f"\n{'═' * 60}")
    print(f"  Concluído! {processados} fornecedor(es) processado(s).")
    print(f"  Saída: {output_file}")
    print(f"{'═' * 60}")


if __name__ == "__main__":
    main()
